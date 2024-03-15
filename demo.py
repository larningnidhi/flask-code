from flask import Flask, render_template, request
from openpyxl import load_workbook
import re 


app = Flask(__name__)

@app.route('/read', methods=['GET', 'POST'])
def read():   
    workbook = load_workbook('data.xlsx')
    worksheet = workbook['Sheet']
    row = worksheet.max_row
    column = worksheet.max_column
    data = []
    for i in range(1, row + 1):
        row_data = []
        for j in range(1, column + 1):
            row_data.append(worksheet.cell(i, j).value)
        data.append(row_data)
    return render_template('p2.html', data=data)

@app.route('/', methods=['GET', 'POST'])
def template_file():
    if request.method == 'POST':

        fname = request.form.get('fname')
        lname = request.form.get('lname')
        if  fname.isdigit() and lname.isdigit():
            errors ="Enter the valid Name"
            return render_template('p1.html',errors=errors)
        
        email = request.form.get('email')
        # pa="^[a-z0-9]@gmail.com"
        # if re.search(pa, email):
        #     pass
        # else:
        #     error_email="Enter the valid Email"
        #     return render_template('p1.html',error_email=error_email)

        city = request.form.get('city')
        date = request.form.get('date')
        age = request.form.get('age')
        gender = request.form.get('gender')
        marks1=int(request.form.get('number1'))
        marks2= int(request.form.get('number2'))
        marks3= int(request.form.get('number3'))
        marks4=int(request.form.get('number4'))
        marks5=int(request.form.get('number5'))
        total=marks1+marks2+marks3+marks4+marks5
        
        
        workbook = load_workbook('data.xlsx')
        worksheet = workbook.active
        
        row = worksheet.max_row + 1
        data_values = [fname, lname, email, city, date, age, gender,marks1,marks2,marks3,marks4,marks5,total]
        for col, value in enumerate(data_values, start=1):
            worksheet.cell(row=row, column=col, value=value)
        workbook.save('data.xlsx')
        
        return read()
        
    else:
        cities = ["gandhinagar", "Ahemdabad", "Gota", "bopal", "puna"]
        return render_template('p1.html', cities=cities)



     


if __name__ == '__main__':  
    app.run(debug=True)